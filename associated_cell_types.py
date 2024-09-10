# -*- coding: utf-8 -*-
# @Author: Xue Chao
# @Time: 2024/03/25 17:51
# @Function: Code for analyses in paper.
import os
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap
from mpl_toolkits.axes_grid1 import make_axes_locatable
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from openpyxl.reader.excel import load_workbook

from util import make_dir, cluster_df
from para import output_DIR, transcriptome_meta_xlsx, GWAS_meta_xlsx, UNSELECT_PHENOS, CELL_CATE_colors


def __heatmap_color_rstyle_single():
    colors = ['#FFFFFF', '#FE0100']
    cmap = LinearSegmentedColormap.from_list("heatmap_rstyle", list(zip([0, 1], colors)))
    return cmap


def __category_color_bar(arr_1d: np.ndarray):
    arr_colors = __color_bar(arr_1d)
    cmap = plt.cm.colors.ListedColormap(arr_colors)
    bounds = np.arange(0, len(arr_1d) + 1)
    norm = plt.cm.colors.BoundaryNorm(bounds, cmap.N)
    return plt.cm.ScalarMappable(norm=norm, cmap=cmap), bounds, arr_colors


def __color_bar(arr, cm='hls'):
    # lut = seaborn.color_palette(cm, len(set(arr)))
    lut = CELL_CATE_colors
    pat = dict(zip(sorted(set(arr)), lut))
    rcos = []
    for co in arr:
        rcos.append(pat[co])
    return rcos


def __df_true_indexes(df: pd.DataFrame):
    xs, ys = [], []
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            if df.iloc[i, j]:
                x, y = j, df.shape[0] - i - 1
                xs.append(x)
                ys.append(y)
    return xs, ys


def heatmap_annot2_rev(color_df: pd.DataFrame, annot1_df: pd.DataFrame, annot2_df: pd.DataFrame,
                       col_colors: [], row_colors: [], area_coor, annot1_label, annot2_label,
                       save_fig=None, max_x_label_cha_len=25, max_abs_r=6):
    area_coor_rev = []
    for a in area_coor:
        area_coor_rev.append([a[1], a[0]])
    heatmap_annot2(color_df.T, annot1_df.T, annot2_df.T, row_colors, col_colors, area_coor_rev, annot1_label,
                   annot2_label, save_fig, max_x_label_cha_len, max_abs_r)


def __add_space(iter, max_cha_len=25):
    arr = []
    for x in iter:
        nx = x
        if len(x) < max_cha_len:
            nx = ' ' * (max_cha_len - len(x)) + x
        arr.append(nx)
    return arr


def heatmap_annot2(color_df: pd.DataFrame, annot1_df: pd.DataFrame, annot2_df: pd.DataFrame,
                   col_colors: [], row_colors: [], area_coor_raw: [], annot1_label, annot2_label,
                   save_fig=None, max_x_label_cha_len=25, max_abs_r=6, color_title=r'$-log_{10}(adjusted\,P)$',
                   annot_legend_show=True, legend_show=True, y_rev_title='', x_rev_title=''):
    cmap = __heatmap_color_rstyle_single()
    scatter_point_base = 180
    fig_size_base = 5 / 15
    nrow, ncol = color_df.shape
    annot1_df = annot1_df.loc[color_df.index, color_df.columns]
    annot2_df = annot2_df.loc[color_df.index, color_df.columns]
    xs = []
    ys = []
    # sizes=[]
    colors = []
    fig, ax = plt.subplots(figsize=(ncol * fig_size_base, nrow * fig_size_base))
    for i in range(nrow):
        for j in range(ncol):
            x, y = j, nrow - i - 1
            ys.append(y)
            xs.append(x)
            colors.append(color_df.iloc[i, j])
    # max_abs_r=color_df.abs().max().max()
    co_ax_map = ax.scatter(xs, ys, c=colors, s=scatter_point_base, alpha=1, cmap=cmap, marker='s', vmin=0,
                           vmax=max_abs_r)
    a1_xs, a1_ys = __df_true_indexes(annot1_df)
    a2_xs, a2_ys = __df_true_indexes(annot2_df)
    ax.scatter(a1_xs, a1_ys, facecolors='none', edgecolors='black', s=scatter_point_base * 0.4, marker='s', alpha=0.7,
               label=annot1_label)
    ax.scatter(a2_xs, a2_ys, facecolors='black', s=scatter_point_base * 0.15, marker='x', alpha=0.7, linewidths=1,
               label=annot2_label)
    ax.set_xlabel(x_rev_title, fontweight='bold', fontsize=14)
    ax.set_ylabel(y_rev_title, fontweight='bold', fontsize=14)
    ax.xaxis.set_label_position('top')
    ax.yaxis.set_label_position('right')

    ## trans df coor to plot coor
    if area_coor_raw is not None:
        area_coor = [[-0.5, nrow - 0.5]]
        for coor in area_coor_raw:
            area_coor.append([color_df.columns.get_loc(coor[1]) - 0.5, nrow - color_df.index.get_loc(coor[0]) - 0.5])
        area_coor.append([ncol - 0.5, -0.5])

        for i in range(1, len(area_coor) - 1):
            nx, ny = area_coor[i]
            nx_b1, ny_b1 = area_coor[i - 1]
            nx_a1, ny_a1 = area_coor[i + 1]
            ax.plot([nx_b1, nx_a1], [ny, ny], '--', linewidth=1.5, c='black')
            ax.plot([nx, nx], [ny_b1, ny_a1], '--', linewidth=1.5, c='black')

    for i in range(nrow - 1):
        ax.axhline(y=0.5 + i, c='black', linewidth=0.05)
    for i in range(ncol - 1):
        ax.axvline(x=0.5 + i, c='black', linewidth=0.05)

    ax.set_aspect('equal')
    ax.set_xlim(-0.5, ncol - 0.5)
    ax.set_ylim(-0.5, nrow - 0.5)
    ax.set_xticks(np.arange(0, color_df.shape[1]), __add_space(color_df.columns, max_x_label_cha_len), rotation=90,
                  ha='center')
    ax.set_yticks(np.arange(0, color_df.shape[0]), color_df.index[::-1])
    ax.tick_params(axis='x', pad=2)
    ax.tick_params(axis='y', pad=2)

    divider = make_axes_locatable(ax)

    if not (col_colors is None or row_colors is None):
        colmap, bound, arr_colors = __category_color_bar(col_colors)
        cax_top = divider.append_axes("top", size=0.1, pad=0.05)
        cax_top_bar = plt.colorbar(colmap, boundaries=bound, cax=cax_top, orientation='horizontal')
        cax_top.xaxis.set_ticks_position('top')

        colmap, bound, arr_colors = __category_color_bar(row_colors[::-1])
        cax_right = divider.append_axes("right", size=0.1, pad=0.05)
        cax_right_bar = plt.colorbar(colmap, boundaries=bound, cax=cax_right)
        cax_right.yaxis.set_ticks_position('right')

        for cbar in [cax_top_bar, cax_right_bar]:
            cbar.set_ticks([])
            cbar.ax.tick_params(size=0)
            cbar.ax.tick_params(which='both', length=0)
            cbar.outline.set_visible(False)

    # b=ax.get_position()
    # x0,x1,y0,y1,w,h=b.x0,b.x1,b.y0,b.y1,b.width,b.height
    # cax_main = fig.add_axes([x0+w/2*1.2,y1-h*0.05,w/2*0.7,0.04*h])
    # cax_main = fig.add_subplot(100,2,2)
    if legend_show:
        cax_main = divider.append_axes('top', size=0.1, pad=0.5)
        cax_main_left = inset_axes(cax_main, width="40%", height="100%", loc='upper left')
        cax_main_bar = plt.colorbar(co_ax_map, cax=cax_main_left, orientation='horizontal')
        cax_main_bar.set_label(label=color_title, labelpad=5)
        if annot_legend_show:
            cax_main_right = inset_axes(cax_main, width="40%", height="100%", loc='upper right')
            le = fig.legend(loc='upper right', bbox_to_anchor=(1, 3.3), bbox_transform=cax_main_right.transAxes)
            # le.set_frame_on(False)
            cax_main_right.set_axis_off()
        cax_main_bar.ax.xaxis.set_label_position('top')
        cax_main.set_axis_off()

    # cax_main_left=inset_axes(cax_main, width="40%",height="100%",loc='left')
    # ax.legend(cax=cax_main_left)
    # cax_main.xaxis.set_ticks_position('top')

    plt.tight_layout()
    plt.show()
    if save_fig is not None:
        fig.savefig(save_fig)


def heatmap_two_marker(df1: pd.DataFrame, df2: pd.DataFrame, color_title=r'Jaccard index', x_title='', y_title='',
                       save_fig=None, max_abs_r=1):
    def two_markers():
        from matplotlib.path import Path
        paths = []
        for i in range(2):
            v = i - 0.5
            verts = [
                (v, v),
                (0.5, -0.5),
                (-0.5, 0.5),
                (v, v),
            ]
            codes = [Path.MOVETO,
                     Path.LINETO,
                     Path.LINETO,
                     Path.CLOSEPOLY,
                     ]
            path = Path(verts, codes)
            paths.append(path)
        return paths

    cmap = __heatmap_color_rstyle_single()
    scatter_point_base = 200
    fig_size_base = 6.5 / 15
    nrow, ncol = df1.shape
    df2 = df2.loc[df1.index, df1.columns]
    fig, ax = plt.subplots(figsize=(ncol * fig_size_base, nrow * fig_size_base))
    # max_abs_r=df1.abs().max().max()
    icons = two_markers()
    for data_df, marker in zip([df1, df2], icons):
        # for data_df, marker in zip([df2], icons[1:2]):
        max_ji = 0
        max_idxs = (0, 0)
        xs = []
        ys = []
        colors = []
        for i in range(nrow):
            for j in range(ncol):
                x, y = j, nrow - i - 1
                ys.append(y)
                xs.append(x)
                v = data_df.iloc[i, j]
                colors.append(v)
                if v > max_ji:
                    max_ji = v
                    max_idxs = (i, j)
        i, j = max_idxs
        # print(f'max: {data_df.index[i]}-{data_df.columns[j]}={data_df.iloc[i,j]}')
        print([f'max: {df.index[i]}-{df.columns[j]}={df.iloc[i, j]}' for df in [df1, df2]])
        co_ax_map = ax.scatter(xs, ys, c=colors, s=scatter_point_base, alpha=1, cmap=cmap, marker=marker, vmin=0,
                               vmax=max_abs_r,
                               )

    ax.set_xlabel(x_title, fontweight='bold', fontsize=14)
    ax.set_ylabel(y_title, fontweight='bold', fontsize=14)
    ax.xaxis.set_label_position('top')
    ax.yaxis.set_label_position('right')

    for i in range(nrow - 1):
        ax.axhline(y=0.5 + i, c='black', lw=0.2)
    for i in range(ncol - 1):
        ax.axvline(x=0.5 + i, c='black', lw=0.2)

    ax.set_aspect('equal')
    ax.set_xlim(-0.5, ncol - 0.5)
    ax.set_ylim(-0.5, nrow - 0.5)
    # ax.set_xticks(np.arange(0,df1.shape[1]),__add_space(df1.columns,max_x_label_cha_len),rotation=90,ha='center')
    ax.set_xticks(np.arange(0, df1.shape[1]), df1.columns, rotation=90, ha='center')
    ax.set_yticks(np.arange(0, df1.shape[0]), df1.index[::-1])
    ax.tick_params(axis='x', pad=2)
    ax.tick_params(axis='y', pad=2)
    #
    divider = make_axes_locatable(ax)
    cax_main = divider.append_axes('top', size=0.1, pad=0.5)
    cax_main_left = inset_axes(cax_main, width="40%", height="100%", loc='upper left')
    cax_main_bar = plt.colorbar(co_ax_map, cax=cax_main_left, orientation='horizontal')
    cax_main_bar.set_label(label=color_title, labelpad=5)
    cax_main_bar.ax.xaxis.set_label_position('top')
    cax_main.set_axis_off()

    plt.tight_layout()
    plt.show()
    if save_fig is not None:
        fig.savefig(save_fig)
    plt.close()

    fig, ax = plt.subplots(figsize=(4, 4))
    for label, ic in zip(['All genes', 'Sig. genes'], icons):
        ax.scatter([], [], marker=ic, label=label, facecolor="#FE0100", edgecolors="none")
    ax.legend(markerscale=1.5)
    plt.show()


def __trans_p_float(x, default=1, cutoff=1e-100):
    num = default
    try:
        num = float(x)
        if num < cutoff:
            num = cutoff
    except:
        pass
    return num


def get_cell_type_abbr_cate(return_source=False):
    meta = transcriptome_meta_xlsx
    wb = load_workbook(meta)
    sheet_names = wb.sheetnames
    abbr_sheets = [s for s in sheet_names if s.split('_')[1] == 'name']
    name_abbr = {}
    abbr_cate = {}
    name_source = {}
    for s in abbr_sheets:
        df = pd.read_excel(meta, sheet_name=s)
        for i in df.index:
            name_abbr[df.loc[i, 'raw_no_empty']] = df.loc[i, 'abbr']
            abbr_cate[df.loc[i, 'abbr']] = df.loc[i, 'category_plot']
            name_source[df.loc[i, 'raw_no_empty']] = s.split('_')[0]
    if return_source:
        return name_abbr, abbr_cate, name_source
    else:
        return name_abbr, abbr_cate


def get_gwas_cate():
    meta = GWAS_meta_xlsx
    abbr_sheets = ['DGN_case']
    abbr_cate = {}
    for s in abbr_sheets:
        df = pd.read_excel(meta, sheet_name=s)
        for i in df.index:
            abbr_cate[df.loc[i, 'abbr']] = df.loc[i, 'category_plot']
    return abbr_cate


def get_gwas_meta(cols: []):
    meta = GWAS_meta_xlsx
    abbr_sheets = ['DGN_case']
    abbr_cate = {}
    for s in abbr_sheets:
        df = pd.read_excel(meta, sheet_name=s)
        for i in df.index:
            abbr_cate[df.loc[i, 'abbr']] = [df.loc[i, x] for x in cols]
    return abbr_cate


def get_ecs_sig_genes():
    pass


def combine_assoc_cells(kggsee_prefixs, pheno_tags, cell_abbrs):
    dfs = []
    for i in range(len(kggsee_prefixs)):
        kp = kggsee_prefixs[i]
        cell_path = f'{kp}.celltype.txt'
        cell_df = pd.read_table(cell_path, header=0, index_col=0, skipfooter=1, engine='python')
        cell_df = cell_df[['Adjusted(p)']]
        cell_df.columns = [pheno_tags[i]]
        dfs.append(cell_df)
        i += 1
    df = pd.concat(dfs, axis=1)
    df = df.applymap(lambda x: -np.log10(__trans_p_float(x)))
    df.index = df.index.map(lambda x: cell_abbrs[x] if x in cell_abbrs else x)
    return df


def plot_assoc_cell_heatmap():
    exclude_phenos = UNSELECT_PHENOS
    sig_p_cutoff = 0.05
    heatmap_cmap = __heatmap_color_rstyle_single()
    kggsee_dirs = []
    for trans_type in ['hs_gtex_gse97930_fc', 'tms_global']:  # 'gtexv8_tpm','tms_global',
        kggsee_dirs.append(f'{output_DIR}/{trans_type}/result')
    plot_datas = []
    for result_dir in kggsee_dirs:
        trans_type = os.path.basename(os.path.dirname(result_dir))
        # gene_score_type='degree.k_adjusted.txt.gz' #degree.k_adjusted expr.mean
        gene_score_type = 'expr.mean.txt.gz'
        gene_score_type_abbr = gene_score_type.split('.')[0]
        expr_type = 'expr.mean.txt.gz'
        dgn_kggsee_prefixs = []
        dese_kggsee_prefixs = []
        pheno_tags = []
        for f in os.listdir(result_dir):
            if f.endswith('.celltype.txt') and gene_score_type in f:
                pheno = f.split(gene_score_type)[0]
                if pheno in exclude_phenos:
                    continue
                if pheno not in exclude_phenos:
                    # expr_p=f'{result_dir}/{pheno}{expr_type}.celltype.txt'
                    # if os.path.isfile(expr_p):
                    dgn_kggsee_prefixs.append(f'{result_dir}/{pheno}{gene_score_type}')
                    dese_kggsee_prefixs.append(f'{result_dir}/{pheno}{expr_type}')
                    pheno_tags.append(pheno)
        if len(dgn_kggsee_prefixs) == 0:
            continue
        cell_abbrs, cell_cate = get_cell_type_abbr_cate()
        gwas_cate = get_gwas_cate()
        df = combine_assoc_cells(dgn_kggsee_prefixs, pheno_tags, cell_abbrs)
        # ex_df=combine_assoc_cells(dese_kggsee_prefixs,pheno_tags,cell_abbrs)
        logp_cut = -np.log10(sig_p_cutoff)
        # only significant association.
        remain_col = df.columns[(df > logp_cut).any()]
        top_row = df.idxmax()
        sig_row = df.index[(df > logp_cut).any(axis=1)]
        remain_row = sorted(set(top_row).union(sig_row))
        # df=df.loc[remain_row,:].loc[:,remain_col]
        df = df.loc[remain_row, :]

        ## sort by category and cluster inner category.
        sorted_gwas_cate = ['Brain', 'Immune/Blood', 'Others']
        sorted_cell_cate = ['Brain', 'Immune/Blood', 'Others']
        sorted_cols = []
        sorted_rows = []
        n_x, n_y = df.shape
        area_coor_idx = [[0, 0]]
        for i in range(len(sorted_gwas_cate)):
            sg = sorted_gwas_cate[i]
            sc = sorted_cell_cate[i]
            cols = [c for c in df.columns if gwas_cate[c] == sg]
            rows = [x for x in df.index if cell_cate[x] == sc]
            if len(cols) > 0 and len(rows) > 0:
                sodf = cluster_df(df.loc[:, cols].loc[rows, :])
                sorted_cols += sodf.columns.tolist()
                sorted_rows += sodf.index.tolist()
            else:
                sorted_cols += cols
                sorted_rows += rows
            last_row, last_col = area_coor_idx[-1]
            area_coor_idx.append([len(rows) + last_row, len(cols) + last_col])
        df = df.loc[:, sorted_cols].loc[sorted_rows, :]
        area_coor = []
        for i in range(1, len(area_coor_idx) - 1):
            x, y = area_coor_idx[i]
            if x >= n_x or y >= n_y:
                continue
            area_coor.append([df.index[x], df.columns[y]])

        row_colors = [cell_cate[i] for i in df.index]
        col_colors = [gwas_cate[i] for i in df.columns]
        annot_df1 = df.applymap(lambda x: x > logp_cut)
        annot_df2 = df.copy()
        for i in df.index:
            for j in df.columns:
                annot_val = False
                # if df.loc[i,j]>logp_cut and ex_df.loc[i,j]<logp_cut:
                if df.loc[i, j] == df.loc[:, j].max() and df.loc[i, j] != 0:
                    annot_val = True
                annot_df2.loc[i, j] = annot_val
        # heatmap_annot2(df,annot_df1,annot_df2,col_colors,row_colors,area_coor,'Significant','Unique')
        plot_datas.append([df, annot_df1, annot_df2, col_colors, row_colors, area_coor, 'Significant', 'Top1'])
    max_log_p = max([pld[0].abs().max().max() for pld in plot_datas])
    for pld in plot_datas:
        heatmap_annot2_rev(*pld, max_abs_r=max_log_p)
        plt.show()
        analysis_dir = f'{output_DIR}/{trans_type}/analysis'
        make_dir(analysis_dir)
        df.to_excel(f'{analysis_dir}/{gene_score_type_abbr}.assoc_cell_matrix.xlsx')

    ## plot category legend
    legend_labels = sorted_cell_cate
    cmap, rang, colors = __category_color_bar(legend_labels)
    fig, ax = plt.subplots()
    legend_lines = [plt.Line2D([0], [0], color=c, lw=4) for c in colors]
    legend = ax.legend(legend_lines, legend_labels, loc='center')
    ax.set_xticks([])
    ax.set_yticks([])
    ax.axis('off')
    plt.show()


def plot_assoc_cell_heatmap_dif():
    exclude_phenos = UNSELECT_PHENOS
    sig_p_cutoff = 0.05
    heatmap_cmap = __heatmap_color_rstyle_single()
    kggsee_dirs = []
    for trans_type in ['hs_gtex_gse97930_fc', 'tms_global']:  # 'gtexv8_tpm','tms_global',
        kggsee_dirs.append(f'{output_DIR}/{trans_type}/dif_result_norez')
    plot_datas = []
    for result_dir in kggsee_dirs:
        trans_type = os.path.basename(os.path.dirname(result_dir))
        gene_score_type = 'degree.z_dif.txt.gz'  # degree.k_adjusted expr.mean  degree.z_dif.txt.gz
        # gene_score_type = 'expr.mean.txt.gz'
        expr_type = 'expr.z_dif.txt.gz'
        dgn_kggsee_prefixs = []
        dese_kggsee_prefixs = []
        pheno_tags = []
        for f in os.listdir(result_dir):
            if f.endswith('.celltype.txt') and gene_score_type in f:
                pheno = f.split(gene_score_type)[0]
                if pheno in exclude_phenos:
                    continue
                if pheno not in exclude_phenos:
                    # expr_p=f'{result_dir}/{pheno}{expr_type}.celltype.txt'
                    # if os.path.isfile(expr_p):
                    dgn_kggsee_prefixs.append(f'{result_dir}/{pheno}{gene_score_type}')
                    dese_kggsee_prefixs.append(f'{result_dir}/{pheno}{expr_type}')
                    pheno_tags.append(pheno)
        if len(dgn_kggsee_prefixs) == 0:
            continue
        cell_abbrs, cell_cate = get_cell_type_abbr_cate()
        gwas_cate = get_gwas_cate()
        df = combine_assoc_cells(dgn_kggsee_prefixs, pheno_tags, cell_abbrs)
        # ex_df=combine_assoc_cells(dese_kggsee_prefixs,pheno_tags,cell_abbrs)
        logp_cut = -np.log10(sig_p_cutoff)
        # only significant association.
        remain_col = df.columns[(df > logp_cut).any()]
        top_row = df.idxmax()
        sig_row = df.index[(df > logp_cut).any(axis=1)]
        remain_row = sorted(set(top_row).union(sig_row))
        # df=df.loc[remain_row,:].loc[:,remain_col]
        df = df.loc[remain_row, :]

        ## sort by category and cluster inner category.
        sorted_gwas_cate = ['Brain', 'Immune/Blood', 'Others']
        sorted_cell_cate = ['Brain', 'Immune/Blood', 'Others']
        sorted_cols = []
        sorted_rows = []
        n_x, n_y = df.shape
        area_coor_idx = [[0, 0]]
        for i in range(len(sorted_gwas_cate)):
            sg = sorted_gwas_cate[i]
            sc = sorted_cell_cate[i]
            cols = [c for c in df.columns if gwas_cate[c] == sg]
            rows = [x for x in df.index if cell_cate[x] == sc]
            if len(cols) > 0 and len(rows) > 0:
                sodf = cluster_df(df.loc[:, cols].loc[rows, :])
                sorted_cols += sodf.columns.tolist()
                sorted_rows += sodf.index.tolist()
            else:
                sorted_cols += cols
                sorted_rows += rows
            last_row, last_col = area_coor_idx[-1]
            area_coor_idx.append([len(rows) + last_row, len(cols) + last_col])
        df = df.loc[:, sorted_cols].loc[sorted_rows, :]
        area_coor = []
        for i in range(1, len(area_coor_idx) - 1):
            x, y = area_coor_idx[i]
            if x >= n_x or y >= n_y:
                continue
            area_coor.append([df.index[x], df.columns[y]])

        row_colors = [cell_cate[i] for i in df.index]
        col_colors = [gwas_cate[i] for i in df.columns]
        annot_df1 = df.applymap(lambda x: x > logp_cut)
        annot_df2 = df.copy()
        for i in df.index:
            for j in df.columns:
                annot_val = False
                # if df.loc[i,j]>logp_cut and ex_df.loc[i,j]<logp_cut:
                if df.loc[i, j] == df.loc[:, j].max() and df.loc[i, j] != 0:
                    annot_val = True
                annot_df2.loc[i, j] = annot_val
        # heatmap_annot2(df,annot_df1,annot_df2,col_colors,row_colors,area_coor,'Significant','Unique')
        plot_datas.append([df, annot_df1, annot_df2, col_colors, row_colors, area_coor, 'Significant', 'Top1'])
    max_log_p = max([pld[0].abs().max().max() for pld in plot_datas])
    for pld in plot_datas:
        heatmap_annot2_rev(*pld, max_abs_r=max_log_p)
        plt.show()
        analysis_dir = f'{output_DIR}/{trans_type}/analysis_dif'
        make_dir(analysis_dir)
        df.to_excel(f'{analysis_dir}/assoc_cell_matrix.xlsx')

    ## plot category legend
    legend_labels = sorted_cell_cate
    cmap, rang, colors = __category_color_bar(legend_labels)
    fig, ax = plt.subplots()
    legend_lines = [plt.Line2D([0], [0], color=c, lw=4) for c in colors]
    legend = ax.legend(legend_lines, legend_labels, loc='center')
    ax.set_xticks([])
    ax.set_yticks([])
    ax.axis('off')
    plt.show()


def main():
    # plot dgn result
    plot_assoc_cell_heatmap()
    plot_assoc_cell_heatmap_dif()


if __name__ == '__main__':
    main()
